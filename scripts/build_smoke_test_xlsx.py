from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Smoke Test"

HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", start_color="D9E2F3")
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="1F3A5F")
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PASS_FILL = PatternFill("solid", start_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
SKIP_FILL = PatternFill("solid", start_color="FFEB9C")
BLOCK_FILL = PatternFill("solid", start_color="F4B084")

headers = ["Section", "#", "Step / Check", "Expected Result", "DB Query / Command", "Status", "Notes"]
widths = [22, 5, 55, 50, 60, 12, 35]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 24
ws.freeze_panes = "A2"

# Structure: list of (section, step, expected, command)
# section == "" means continuation row under prior section heading
rows = [
    # 0. Preconditions
    ("0. Preconditions", "App running on Replit (pnpm dev) and web UI loads", "Login page renders without errors", ""),
    ("", "Database reachable on boot", "No DB connection errors in server log", ""),
    ("", "Aircall API ID + Token saved in /settings", "Save succeeds; credentials not re-posted as masked ****", "Invariant #2 — never re-save masked"),
    ("", "Aircall webhook URL registered in Aircall dashboard", "Points at https://<replit-url>/api/aircall/webhook", ""),
    ("", "Webhook subscriptions enabled", "call.ended, call.tagged, call.commented, transcription.created, summary.created", ""),
    ("", "Test agent exists in agents table with aircall_user_id + email populated", "Agent row maps to Aircall user AND has email for SSO match", "SELECT id, name, email, aircall_user_id FROM agents WHERE email IS NOT NULL;"),
    ("", "Real phone available to answer test calls", "You or teammate can pick up", ""),
    ("", "Clear webhook log for clean visibility", "Log empty", "GET /api/aircall/webhook-log?clear=1"),

    # 0.5 Auth preconditions
    ("0.5 Auth Preconditions", "GCP OAuth 2.0 client created (Web application type)", "Client ID and Secret in hand", ""),
    ("", "Authorized JavaScript origins includes APP_URL", "Exact match, no trailing slash", ""),
    ("", "Authorized redirect URI: {APP_URL}/api/auth/callback", "Exact match required by OAuth", ""),
    ("", "Env vars set on Replit Secrets", "GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET, SESSION_SECRET, APP_URL all non-empty", ""),
    ("", "SESSION_SECRET is 32+ bytes random", "Generate with: openssl rand -base64 32", ""),
    ("", "Every test agent row has email populated", "SSO access gate matches by email — rows with NULL email cannot log in", "SELECT id, name, email FROM agents WHERE email IS NULL; -- should return 0 rows"),

    # 1. Sign in flow
    ("1. Sign in (Google SSO)", "Hit app root URL", "Redirects to /login", ""),
    ("", "Click 'Sign in with Google'", "Redirects to accounts.google.com consent", ""),
    ("", "Complete Google consent as an allowed agent (email ∈ agents.email)", "Redirects back to app root; sidebar shows real name + avatar", ""),
    ("", "DB — user row created", "One row in users with matching google_sub + email", "SELECT id, email, google_sub, last_login_at FROM users ORDER BY last_login_at DESC LIMIT 5;"),
    ("", "DB — agent bound to user", "agents.user_id set to the user.id from previous step", "SELECT a.name, a.email, a.user_id, u.email AS user_email FROM agents a LEFT JOIN users u ON u.id = a.user_id;"),
    ("", "GET /api/auth/me returns session", "HTTP 200 with { user, agent } shape", "Open DevTools → Network; refresh; inspect /api/auth/me response"),
    ("", "Logout button signs out cleanly", "Click avatar area → logout icon → lands on /login", ""),
    ("", "Rejection test — sign in with non-agent Google account", "403 Access denied with the rejected email shown", ""),

    # 2. Add Contacts
    ("2. Add Contacts", "Navigate to Contacts (contact-ingestion page)", "Page loads with Upload / Paste / History tabs", ""),
    ("", "Paste 3-row test CSV (1 smoke, 1 duplicate phone, 1 other)", "", "first_name,last_name,phone,email,company\nSmoke,Test,+44YOURNUMBER,smoke@test.local,SmokeCo\nDupe,Check,+44YOURNUMBER,other@test.local,DupeCo\nJane,Valid,+447700900001,jane@test.local,ValidCo"),
    ("", "Click analyse / suggest mapping", "All 5 columns auto-detected", ""),
    ("", "Click stage", "3 rows staged; row 2 flagged duplicate on phone", ""),
    ("", "Review + commit", "Commit reports 2 new, 1 duplicate (or similar)", ""),
    ("", "DB sanity check", "Rows present, dispatch_status='pool', upload_batch populated, no duplicate phones", "SELECT id, first_name, phone, dispatch_status, source_list, upload_batch FROM contacts WHERE phone LIKE '%YOURNUMBER%' OR email LIKE '%test.local';"),

    # 3. Create Call List
    ("3. Create Call List", "Navigate to Call Lists, click New Call List", "Create dialog opens", ""),
    ("", "Fill form: name, daily_quota=10, assigned_agent, source_lists=[upload_batch from step 1]", "Form accepts input", ""),
    ("", "Create list", "List appears; queue status ~2 fresh needed, 0 callbacks/interested/retries", ""),
    ("", "Click Fill Queue", "2 dispatched (or however many test contacts)", ""),
    ("", "DB check — contacts dispatched + membership active", "Both rows active, removed_at IS NULL", "SELECT c.phone, c.dispatch_status, m.call_list_id, m.removed_at FROM contacts c JOIN call_list_memberships m ON m.contact_id=c.id WHERE m.removed_at IS NULL AND c.email LIKE '%test.local';"),

    # 4. Make a Call
    ("4. Make a Call", "Navigate to Call Command", "Page loads", ""),
    ("", "Select your test agent in picker", "Only assigned lists visible; other agents' lists hidden", "Invariant #5 — agent scoping"),
    ("", "Select Smoke Test List", "Queue shows test contact at top; Up Next shows second", ""),
    ("", "Click Dial on Smoke/Test", "Aircall dialer opens, call connects, phone rings", ""),
    ("", "Answer + talk 15+ seconds + hang up", "Call completes cleanly", "Need audio duration for transcription to be meaningful"),
    ("", "Check webhook log for call.ended", "Event present within ~30s of hangup", "GET /api/aircall/webhook-log"),
    ("", "DB check — lead_conversations row written", "external_id = Aircall call ID, contact_id matched, duration_seconds > 0, agent_name correct", "SELECT id, external_id, contact_id, direction, duration_seconds, agent_name, conversation_date FROM lead_conversations ORDER BY conversation_date DESC LIMIT 1;"),

    # 5. Record Outcome
    ("5. Record Outcome", "In Aircall UI, tag the call with Interested (or mapped equivalent)", "Tag saved on Aircall side", ""),
    ("", "Call Command shows pending-outcome pill then refreshes", "SSE or polling fallback fires", ""),
    ("", "Webhook log shows call.tagged event", "Tag extracted correctly (tagModel handles 7 payload shapes)", ""),
    ("", "DB — contact state updated", "last_call_outcome='interested', call_attempts incremented, dispatch_status changed", "SELECT c.phone, c.last_call_outcome, c.dispatch_status, c.call_attempts FROM contacts c WHERE c.email='smoke@test.local';"),
    ("", "DB — membership closed", "removed_at set, removal_reason='called', outcome_at_removal='interested'", "SELECT removed_at, removal_reason, outcome_at_removal FROM call_list_memberships WHERE contact_id=(SELECT id FROM contacts WHERE email='smoke@test.local');"),
    ("", "Outcome drawer displays the outcome", "Drawer shows interested badge + any engine output if engine ran", ""),
    ("", "Today's Results counter ticks up", "Completed / Interested counters increment", ""),
    ("", "Up Next reflows with cursor re-aligned by contact ID", "Operator is not jumped to wrong contact", "Invariant #4"),
    ("", "Repeat for tag: no-answer", "Contact returns to pool with cool-off applied", ""),
    ("", "Repeat for tag: callback-requested", "callback_date set; queue reflects it", ""),
    ("", "Repeat for tag: no-interest", "Contact archived / cooled-off per config", ""),

    # 6. Transcript
    ("6. Transcript Capture", "Wait 2-5 min after hangup", "Aircall AI processes the call", ""),
    ("", "Webhook log shows transcription.created", "Event received", ""),
    ("", "DB — transcript_text populated", "Non-null, speaker labels collapsed cleanly", "SELECT LENGTH(transcript_text) AS t_len, transcript_text FROM lead_conversations WHERE external_id='<Aircall call ID>';"),

    # 7. Summary
    ("7. Summary Capture", "Webhook log shows summary.created", "Event received (newer Aircall tenants fire this separately)", ""),
    ("", "DB — summary populated", "Aircall native AI summary stored", "SELECT summary FROM lead_conversations WHERE external_id='<Aircall call ID>';"),
    ("", "NULL summary acceptable for very short calls", "Not a blocker — no LLM fallback today", ""),

    # 8. Multi-agent via separate Google accounts / browser profiles
    ("8. Multi-Agent Isolation", "Prerequisite: 2+ agents in DB, each with email matching a real Google Workspace account", "", ""),
    ("", "Agent A signs in on Machine 1 (or Chrome profile 1)", "Sees only their assigned + unassigned lists", ""),
    ("", "Agent B signs in on Machine 2 (or Chrome profile 2)", "Sees only their own assigned + unassigned lists", ""),
    ("", "Today's Results, stale count, queue — all scoped to the logged-in agent", "No cross-agent leakage on either browser", "Invariant #5 — now server-enforced via req.auth"),
    ("", "Agent A makes a call → Agent B shouldn't see it in Today's Results", "Each agent's outcomes tray reflects only their calls", ""),
    ("", "Agent A hits /api/call-lists without cookie (e.g. curl)", "401 not_authenticated", "curl -i https://{APP_URL}/api/call-lists"),
    ("", "Agent A swaps ?agent_id=AGENT_B_ID in a URL — no effect", "Endpoint ignores client agent_id; scope from session only", ""),

    # 8.5 Aircall mismatch
    ("8.5 Aircall Identity Mismatch", "Sign into app as Agent A, sign into Aircall widget as Agent A", "No mismatch banner shown", ""),
    ("", "Log out of Aircall widget, log in as Agent B (while still app-logged as A)", "Amber banner appears on Call Command: 'Aircall is signed in as a different user'", ""),
    ("", "Log out of Aircall again, log back in as Agent A", "Banner clears", ""),
    ("", "Banner does NOT block dialling — it's advisory", "Dial button still works when banner visible", "Attribution will follow the Aircall user, not the app user — acceptable, surfaced loudly"),

    # 9. Cleanup
    ("9. Cleanup + Sign-off", "Archive or delete Smoke Test List", "List no longer active", ""),
    ("", "Mark test contacts dispatch_status='archived' or delete", "Test data cleared", ""),
    ("", "Clear webhook log", "Log empty for next run", "GET /api/aircall/webhook-log?clear=1"),
    ("", "All 6 functions pass end-to-end", "No errors in server log during run", ""),
    ("", "Signed off — ready to hand to first agent", "Date + operator name noted", ""),
]

current_row = 2
for section, step, expected, command in rows:
    if section:
        # New section heading spanning first two cols
        ws.cell(row=current_row, column=1, value=section).font = SECTION_FONT
        for c in range(1, 8):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = SECTION_FILL
            cell.border = BORDER
            cell.alignment = WRAP
    # Step row
    ws.cell(row=current_row, column=2, value=current_row - 1).alignment = CENTER
    ws.cell(row=current_row, column=3, value=step).alignment = WRAP
    ws.cell(row=current_row, column=4, value=expected).alignment = WRAP
    ws.cell(row=current_row, column=5, value=command).alignment = WRAP
    ws.cell(row=current_row, column=6, value="Pending").alignment = CENTER
    ws.cell(row=current_row, column=7, value="").alignment = WRAP
    for c in range(1, 8):
        cell = ws.cell(row=current_row, column=c)
        if not cell.font or cell.font.name != "Arial":
            cell.font = BODY_FONT
        cell.border = BORDER
    # Row height — enough for wrapped text
    ws.row_dimensions[current_row].height = max(22, 15 * (1 + (len(step) // 55) + (len(expected) // 50) + (len(command) // 60)))
    current_row += 1

# Data validation for Status column
dv = DataValidation(type="list", formula1='"Pass,Fail,Skip,Blocked,Pending"', allow_blank=True)
dv.add(f"F2:F{current_row - 1}")
ws.add_data_validation(dv)

# Conditional formatting on Status column
status_range = f"F2:F{current_row - 1}"
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL))
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL))
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Skip"'], fill=SKIP_FILL))
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_FILL))

# Summary sheet
summary = wb.create_sheet("Summary")
summary.column_dimensions["A"].width = 28
summary.column_dimensions["B"].width = 14

summary["A1"] = "MVP Smoke Test — Summary"
summary["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3A5F")
summary.merge_cells("A1:B1")

summary["A3"] = "Total steps"
summary["B3"] = f"=COUNTA('Smoke Test'!C2:C{current_row - 1})"
summary["A4"] = "Passed"
summary["B4"] = f'=COUNTIF(\'Smoke Test\'!F2:F{current_row - 1},"Pass")'
summary["A5"] = "Failed"
summary["B5"] = f'=COUNTIF(\'Smoke Test\'!F2:F{current_row - 1},"Fail")'
summary["A6"] = "Skipped"
summary["B6"] = f'=COUNTIF(\'Smoke Test\'!F2:F{current_row - 1},"Skip")'
summary["A7"] = "Blocked"
summary["B7"] = f'=COUNTIF(\'Smoke Test\'!F2:F{current_row - 1},"Blocked")'
summary["A8"] = "Pending"
summary["B8"] = f'=COUNTIF(\'Smoke Test\'!F2:F{current_row - 1},"Pending")'
summary["A9"] = "% complete"
summary["B9"] = f"=(B4+B5+B6+B7)/B3"
summary["B9"].number_format = "0.0%"

summary["A11"] = "Operator"
summary["A12"] = "Date"
summary["A13"] = "Git commit / branch"
summary["A14"] = "Overall verdict (Ready to Ship / Needs Fix)"

for row in range(3, 10):
    summary.cell(row=row, column=1).font = BODY_FONT
    summary.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True)
for row in range(11, 15):
    summary.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True)
    summary.cell(row=row, column=2).fill = PatternFill("solid", start_color="FFF2CC")
    summary.cell(row=row, column=2).border = BORDER

wb.save(r"C:\Users\Werner\Documents\Apps\Work\content-unlocker\content-unlocker\MVP_Smoke_Test.xlsx")
print("OK")
